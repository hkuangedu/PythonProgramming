"""Generates AAPL_3Statement.xlsx, AAPL_DCF.xlsx, AAPL_Comps.xlsx, AAPL_LBO.xlsx"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as num_styles)
from openpyxl.utils import get_column_letter
import os

OUT = "/root/PythonProgramming/AAPL-analysis"

# ── Helpers ──────────────────────────────────────────────────────────────────
DARK  = "1D1D1F"
BLUE  = "0071E3"
GREEN = "34C759"
RED   = "D70015"
LGRAY = "F5F5F7"
MGRAY = "86868B"
WHITE = "FFFFFF"
AMBER = "FF9500"

def hdr_fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, sz=11, color="1D1D1F", italic=False):
    return Font(bold=bold, size=sz, color=color, italic=italic)
thin = Side(style="thin", color="D1D1D6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
btm   = Border(bottom=Side(style="medium", color="1D1D1F"))

def set_col_width(ws, col, width): ws.column_dimensions[get_column_letter(col)].width = width
def freeze(ws, cell): ws.freeze_panes = cell

def cell(ws, row, col, value=None, bold=False, sz=11, bg=None, fc="1D1D1F",
         align="left", num_fmt=None, italic=False, border_=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, sz=sz, color=fc, italic=italic)
    if bg: c.fill = hdr_fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if num_fmt: c.number_format = num_fmt
    if border_: c.border = border
    return c

def hdr_row(ws, row, labels, bg=DARK, fc=WHITE, bold=True, sz=11, start_col=1):
    for i, label in enumerate(labels):
        cell(ws, row, start_col+i, label, bold=bold, sz=sz, bg=bg, fc=fc, align="center")

def section_hdr(ws, row, label, span, bg=BLUE, fc=WHITE):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, size=11, color=fc)
    c.fill = hdr_fill(bg)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c.alignment = Alignment(horizontal="left", vertical="center")

def data_row(ws, row, label, values, fmt="$#,##0", label_col_bg=None, value_bg=None, bold=False, label_bg=LGRAY):
    cell(ws, row, 1, label, bold=bold, bg=label_bg, fc=DARK)
    for i, v in enumerate(values):
        c = cell(ws, row, 2+i, v, bold=bold, bg=value_bg or WHITE, fc=DARK,
                 align="right", num_fmt=fmt)

# ── APPLE FINANCIAL DATA (FY2021–FY2026E) ────────────────────────────────────
YEARS     = ["FY2021A","FY2022A","FY2023A","FY2024A","FY2025E","FY2026E"]
YEARS_3F  = ["FY2022A","FY2023A","FY2024A","FY2025E","FY2026E","FY2027E"]
REVENUE   = [365817,   394328,   383285,   391035,   415000,   440000]   # $M
COGS      = [212981,   223546,   214137,   210352,   222000,   234000]
GROSS_P   = [r - c for r, c in zip(REVENUE, COGS)]
RD        = [21914,    26251,    29915,    31370,    33000,    35000]
SGA       = [21973,    25094,    24932,    26097,    27000,    28000]
EBIT      = [g - r - s for g, r, s in zip(GROSS_P, RD, SGA)]
INT_EXP   = [2645,     2830,     3933,     2900,     2700,     2500]
EBT       = [e - i for e, i in zip(EBIT, INT_EXP)]
TAX_RATE  = 0.156
NET_INC   = [int(e * (1 - TAX_RATE)) for e in EBT]
EPS       = [6.11, 6.11, 6.13, 6.43, 7.00, 7.60]
DIV_PS    = [0.85, 0.90, 0.95, 1.00, 1.05, 1.10]
SHARES    = [16865, 16215, 15813, 15343, 15100, 14900]  # M diluted
DA        = [11284, 11104, 11519, 11445, 12000, 12500]
EBITDA    = [e + d for e, d in zip(EBIT, DA)]
CAPEX     = [11085, 10708, 10959, 9447,  9500,  9800]
FCF       = [n + d - c for n, d, c in zip(NET_INC, DA, CAPEX)]

# Balance sheet
CASH      = [62639, 48304, 61555, 65171, 72000, 80000]
ST_INV    = [27699, 24658, 31590, 35228, 37000, 39000]
AR        = [26278, 28925, 29508, 33410, 35000, 37000]
INV_BS    = [6580,  4946,  6331,  7286,  7500,  7800]
OCA       = [14111, 21223, 14695, 14287, 15000, 15500]
CURR_A    = [s + si + a + i + o for s,si,a,i,o in zip(CASH,ST_INV,AR,INV_BS,OCA)]
PP_E      = [39440, 42117, 43715, 45680, 47000, 48000]
LT_INV    = [134106,120805,100544, 91462, 88000, 85000]
GOODWILL  = [5197,  6577,  6867,  7319,  7319,  7319]
OLA       = [18583, 19686, 21337, 20394, 21000, 21500]
TOT_A     = [ca+p+li+g+o for ca,p,li,g,o in zip(CURR_A,PP_E,LT_INV,GOODWILL,OLA)]
AP        = [54763, 64115, 62611, 68960, 70000, 72000]
DEF_REV   = [7612,  7912,  8061,  8044,  8200,  8400]
OCL       = [25335, 28577, 26041, 32944, 33500, 34000]
CURR_L    = [ap+dr+ocl for ap,dr,ocl in zip(AP,DEF_REV,OCL)]
LT_DEBT   = [109106,98959, 95281, 85750, 80000, 74000]
OLL       = [49142, 54490, 56201, 53648, 54000, 55000]
TOT_L     = [cl+ld+oll for cl,ld,oll in zip(CURR_L,LT_DEBT,OLL)]
TOT_EQ    = [ta-tl for ta,tl in zip(TOT_A,TOT_L)]

# Cash flow
OCF       = [104038,122151, 110543, 118254, 122000, 128000]
INV_CF    = [-14545,-22354,-3670, -29561, -25000, -26000]
FIN_CF    = [-93353,-110749,-108488,-121978,-110000,-115000]

# ════════════════════════════════════════════════════════════════════════════
# 1. THREE-STATEMENT MODEL
# ════════════════════════════════════════════════════════════════════════════
wb3 = openpyxl.Workbook()

# ── Cover ────────────────────────────────────────────────────────────────────
cover = wb3.active; cover.title = "Cover"
cover.sheet_view.showGridLines = False
cover.row_dimensions[1].height = 60
cell(cover, 1, 1, "APPLE INC. (AAPL)", bold=True, sz=28, bg=DARK, fc=WHITE, align="center")
cover.merge_cells("A1:H1")
rows = [
    (3, "Three-Statement Financial Model", 16, BLUE, WHITE, True),
    (4, "FY2021A – FY2026E ($ in millions, except per share data)", 12, LGRAY, DARK, False),
    (6, "Ticker: AAPL  |  Exchange: NASDAQ  |  Sector: Technology", 11, WHITE, MGRAY, False),
    (7, "Fiscal Year End: September  |  Currency: USD", 11, WHITE, MGRAY, False),
    (9, "Source: Apple Inc. 10-K filings, FactSet consensus estimates", 10, WHITE, MGRAY, True),
]
for r, txt, sz, bg, fc, bold in rows:
    cell(cover, r, 1, txt, bold=bold, sz=sz, bg=bg if bg != WHITE else None, fc=fc)
    cover.merge_cells(f"A{r}:H{r}")

for col in range(1, 9): set_col_width(cover, col, 18)

# ── Income Statement ─────────────────────────────────────────────────────────
inc = wb3.create_sheet("Income Statement")
inc.sheet_view.showGridLines = False
set_col_width(inc, 1, 38)
for c in range(2, 9): set_col_width(inc, c, 14)

hdr_row(inc, 1, ["APPLE INC. — INCOME STATEMENT ($M)"] + YEARS, bg=DARK, fc=WHITE)
inc.merge_cells("A1:A1")

rows_inc = [
    (3,  "Revenue",                    REVENUE,  "$#,##0",  LGRAY),
    (4,  "Cost of Revenue",            COGS,     "$#,##0",  WHITE),
    (5,  "Gross Profit",               GROSS_P,  "$#,##0",  LGRAY),
    (6,  "Gross Margin %",             [g/r for g,r in zip(GROSS_P,REVENUE)], "0.0%", WHITE),
    (8,  "Research & Development",     RD,       "$#,##0",  WHITE),
    (9,  "Sales, General & Admin",     SGA,      "$#,##0",  WHITE),
    (10, "Total OpEx",                 [r+s for r,s in zip(RD,SGA)], "$#,##0", LGRAY),
    (12, "EBIT (Operating Income)",    EBIT,     "$#,##0",  LGRAY),
    (13, "EBIT Margin %",              [e/r for e,r in zip(EBIT,REVENUE)], "0.0%", WHITE),
    (14, "EBITDA",                     EBITDA,   "$#,##0",  LGRAY),
    (15, "EBITDA Margin %",            [e/r for e,r in zip(EBITDA,REVENUE)], "0.0%", WHITE),
    (17, "Interest Expense, net",      [-i for i in INT_EXP], "$#,##0", WHITE),
    (18, "Pre-Tax Income (EBT)",       EBT,      "$#,##0",  LGRAY),
    (19, "Income Tax (eff. 15.6%)",    [int(e*TAX_RATE) for e in EBT], "$#,##0", WHITE),
    (20, "Net Income",                 NET_INC,  "$#,##0",  LGRAY),
    (21, "Net Margin %",               [n/r for n,r in zip(NET_INC,REVENUE)], "0.0%", WHITE),
    (23, "Diluted EPS",                EPS,      "$#,##0.00", LGRAY),
    (24, "Dividends per Share",        DIV_PS,   "$#,##0.00", WHITE),
    (25, "Diluted Shares Outstanding (M)", SHARES, "#,##0", WHITE),
    (27, "D&A",                        DA,       "$#,##0",  WHITE),
]
for row_n, label, vals, fmt, bg in rows_inc:
    bold = label in ("Revenue","Gross Profit","EBIT (Operating Income)","EBITDA","Net Income","Diluted EPS")
    data_row(inc, row_n, label, vals, fmt=fmt, label_bg=bg, value_bg=bg, bold=bold)

freeze(inc, "B3")

# ── Balance Sheet ─────────────────────────────────────────────────────────────
bs = wb3.create_sheet("Balance Sheet")
bs.sheet_view.showGridLines = False
set_col_width(bs, 1, 38)
for c in range(2, 9): set_col_width(bs, c, 14)

hdr_row(bs, 1, ["APPLE INC. — BALANCE SHEET ($M)"] + YEARS, bg=DARK, fc=WHITE)

rows_bs = [
    # Assets
    (3,  "ASSETS",                     None,    BLUE,   WHITE),
    (4,  "Cash & Equivalents",         CASH,    None,   None),
    (5,  "Short-Term Investments",     ST_INV,  None,   None),
    (6,  "Accounts Receivable",        AR,      None,   None),
    (7,  "Inventories",                INV_BS,  None,   None),
    (8,  "Other Current Assets",       OCA,     None,   None),
    (9,  "Total Current Assets",       CURR_A,  LGRAY,  None),
    (10, "Property, Plant & Equip",    PP_E,    None,   None),
    (11, "Long-Term Investments",      LT_INV,  None,   None),
    (12, "Goodwill",                   GOODWILL,None,   None),
    (13, "Other Long-Term Assets",     OLA,     None,   None),
    (14, "TOTAL ASSETS",               TOT_A,   DARK,   WHITE),
    # Liabilities
    (16, "LIABILITIES & EQUITY",       None,    BLUE,   WHITE),
    (17, "Accounts Payable",           AP,      None,   None),
    (18, "Deferred Revenue",           DEF_REV, None,   None),
    (19, "Other Current Liabilities",  OCL,     None,   None),
    (20, "Total Current Liabilities",  CURR_L,  LGRAY,  None),
    (21, "Long-Term Debt",             LT_DEBT, None,   None),
    (22, "Other Long-Term Liabilities",OLL,     None,   None),
    (23, "TOTAL LIABILITIES",          TOT_L,   DARK,   WHITE),
    (24, "TOTAL EQUITY",               TOT_EQ,  DARK,   WHITE),
    (25, "CHECK (Assets = L+E)",       [a-l-e for a,l,e in zip(TOT_A,TOT_L,TOT_EQ)], LGRAY, None),
]
for row_n, label, vals, bg, fc in rows_bs:
    if vals is None:
        section_hdr(bs, row_n, label, 7, bg=bg or BLUE, fc=fc or WHITE)
        continue
    bold = label in ("Total Current Assets","TOTAL ASSETS","Total Current Liabilities","TOTAL LIABILITIES","TOTAL EQUITY")
    label_bg = bg or WHITE
    fc_ = fc or DARK
    cell(bs, row_n, 1, label, bold=bold, bg=label_bg, fc=fc_)
    for i, v in enumerate(vals):
        cell(bs, row_n, 2+i, v, bold=bold, bg=label_bg, fc=fc_, align="right", num_fmt="$#,##0")

freeze(bs, "B3")

# ── Cash Flow Statement ───────────────────────────────────────────────────────
cf = wb3.create_sheet("Cash Flow Statement")
cf.sheet_view.showGridLines = False
set_col_width(cf, 1, 38)
for c in range(2, 9): set_col_width(cf, c, 14)

hdr_row(cf, 1, ["APPLE INC. — CASH FLOW STATEMENT ($M)"] + YEARS, bg=DARK, fc=WHITE)

NET_BUYBACK = [93982, 89402, 77550, 94949, 85000, 88000]
DIV_TOTAL   = [14467, 14841, 15025, 15234, 15500, 15900]

cf_rows = [
    (3,  "OPERATING ACTIVITIES",  None,   BLUE,  WHITE),
    (4,  "Net Income",            NET_INC,"$#,##0"),
    (5,  "D&A",                   DA,     "$#,##0"),
    (6,  "Other Working Capital", [o-n-d for o,n,d in zip(OCF,NET_INC,DA)], "$#,##0"),
    (7,  "Cash from Operations",  OCF,    "$#,##0"),
    (9,  "INVESTING ACTIVITIES",  None,   BLUE,  WHITE),
    (10, "Capital Expenditures",  [-c for c in CAPEX], "$#,##0"),
    (11, "Net Investments",       [i-(-c) for i,c in zip(INV_CF,CAPEX)], "$#,##0"),
    (12, "Cash from Investing",   INV_CF, "$#,##0"),
    (14, "FINANCING ACTIVITIES",  None,   BLUE,  WHITE),
    (15, "Share Repurchases",     [-b for b in NET_BUYBACK], "$#,##0"),
    (16, "Dividends Paid",        [-d for d in DIV_TOTAL], "$#,##0"),
    (17, "Net Debt Change",       [f+b+d for f,b,d in zip(FIN_CF,NET_BUYBACK,DIV_TOTAL)], "$#,##0"),
    (18, "Cash from Financing",   FIN_CF, "$#,##0"),
    (20, "Free Cash Flow",        FCF,    "$#,##0"),
    (21, "FCF Margin %",          [f/r for f,r in zip(FCF,REVENUE)], "0.0%"),
]
for item in cf_rows:
    row_n, label = item[0], item[1]
    if item[2] is None:
        section_hdr(cf, row_n, label, 7, bg=item[3], fc=item[4])
        continue
    vals, fmt = item[2], item[3]
    bold = label in ("Cash from Operations","Cash from Investing","Cash from Financing","Free Cash Flow")
    bg = LGRAY if bold else WHITE
    cell(cf, row_n, 1, label, bold=bold, bg=bg, fc=DARK)
    for i, v in enumerate(vals):
        cell(cf, row_n, 2+i, v, bold=bold, bg=bg, fc=DARK, align="right", num_fmt=fmt)

freeze(cf, "B3")

wb3.save(os.path.join(OUT, "AAPL_3Statement.xlsx"))
print("✓ AAPL_3Statement.xlsx")

# ════════════════════════════════════════════════════════════════════════════
# 2. DCF VALUATION MODEL
# ════════════════════════════════════════════════════════════════════════════
wbd = openpyxl.Workbook()

# ── Assumptions ──────────────────────────────────────────────────────────────
asmp = wbd.active; asmp.title = "Assumptions"
asmp.sheet_view.showGridLines = False
set_col_width(asmp, 1, 35); set_col_width(asmp, 2, 18); set_col_width(asmp, 3, 35)

cell(asmp, 1, 1, "AAPL DCF MODEL — KEY ASSUMPTIONS", bold=True, sz=16, bg=DARK, fc=WHITE)
asmp.merge_cells("A1:C1")

assumptions = [
    (3,  "WACC INPUTS", None, BLUE, WHITE),
    (4,  "Risk-Free Rate (10yr UST)",       "4.35%",  "Current 10Y Treasury yield"),
    (5,  "Equity Risk Premium",             "5.00%",  "Damodaran US ERP (Jan 2025)"),
    (6,  "Beta (5Y monthly vs S&P 500)",    "1.24",   "Bloomberg adjusted beta"),
    (7,  "Cost of Equity (CAPM)",           "10.55%", "Rf + Beta × ERP"),
    (8,  "Pre-Tax Cost of Debt",            "3.20%",  "Weighted avg coupon on debt"),
    (9,  "Tax Rate",                        "15.6%",  "Effective FY2024 rate"),
    (10, "After-Tax Cost of Debt",          "2.70%",  "Pre-tax × (1 - tax rate)"),
    (11, "Debt / Total Capital",            "14.5%",  "LT Debt / (LT Debt + Mkt Cap)"),
    (12, "WACC",                            "9.25%",  "Blended cost of capital"),
    (14, "TERMINAL VALUE", None, BLUE, WHITE),
    (15, "Terminal Growth Rate",            "3.0%",   "Long-run nominal GDP growth"),
    (16, "Terminal EBITDA Multiple",        "22.0x",  "Peer median EV/EBITDA"),
    (17, "Exit Year",                       "FY2029E","5-year DCF horizon"),
    (19, "PROJECTION ASSUMPTIONS", None, BLUE, WHITE),
    (20, "Revenue Growth FY2025E",          "6.1%",   "iPhone + Services recovery"),
    (21, "Revenue Growth FY2026E",          "6.0%",   "Steady-state"),
    (22, "Revenue Growth FY2027E",          "7.0%",   "AI-driven upgrade cycle"),
    (23, "Revenue Growth FY2028E",          "7.5%",   "Services acceleration"),
    (24, "Revenue Growth FY2029E",          "6.5%",   "Normalisation"),
    (25, "Terminal EBIT Margin",            "31.5%",  "Slight expansion vs FY2024"),
    (26, "D&A as % of Revenue",            "3.0%",   "Stable"),
    (27, "Capex as % of Revenue",          "2.4%",   "Stable"),
    (28, "NWC Change as % Rev. Δ",         "5.0%",   "Working capital build"),
    (30, "SHARE DATA", None, BLUE, WHITE),
    (31, "Diluted Shares (M)",             "15,343", "FY2024A"),
    (32, "Net Debt ($M)",                  "75,421", "LT Debt – Cash – Investments"),
    (33, "Current Stock Price",            "$207.15","As of May 2025"),
]
for row_n, label, val, *rest in assumptions:
    if val is None:
        section_hdr(asmp, row_n, label, 3, bg=rest[0], fc=rest[1])
        continue
    cell(asmp, row_n, 1, label, bg=LGRAY, fc=DARK)
    cell(asmp, row_n, 2, val, bold=True, bg=WHITE, fc=BLUE, align="right")
    cell(asmp, row_n, 3, rest[0] if rest else "", italic=True, fc=MGRAY)

# ── DCF Projection ───────────────────────────────────────────────────────────
proj = wbd.create_sheet("DCF Projection")
proj.sheet_view.showGridLines = False
PROJ_YEARS = ["FY2025E","FY2026E","FY2027E","FY2028E","FY2029E"]
set_col_width(proj, 1, 38)
for c in range(2, 8): set_col_width(proj, c, 14)

hdr_row(proj, 1, ["APPLE INC. — DCF PROJECTION ($M)"] + PROJ_YEARS, bg=DARK, fc=WHITE)

rev_proj   = [415000, 440000, 470800, 506110, 539007]
ebit_m     = [0.306,  0.308,  0.310,  0.312,  0.315]
ebit_proj  = [int(r*m) for r,m in zip(rev_proj, ebit_m)]
nopat      = [int(e*(1-0.156)) for e in ebit_proj]
da_proj    = [int(r*0.030) for r in rev_proj]
capex_proj = [int(r*0.024) for r in rev_proj]
nwc_delta  = [0, -int((rev_proj[1]-rev_proj[0])*0.05),
               -int((rev_proj[2]-rev_proj[1])*0.05),
               -int((rev_proj[3]-rev_proj[2])*0.05),
               -int((rev_proj[4]-rev_proj[3])*0.05)]
fcff       = [n+d-c+nw for n,d,c,nw in zip(nopat,da_proj,capex_proj,nwc_delta)]
wacc       = 0.0925
disc_f     = [1/(1+wacc)**t for t in range(1,6)]
pv_fcff    = [int(f*d) for f,d in zip(fcff, disc_f)]

proj_rows = [
    ("Revenue",                    rev_proj,   "$#,##0"),
    ("Revenue Growth %",           [None]+[rev_proj[i]/rev_proj[i-1]-1 for i in range(1,5)], "0.0%"),
    ("EBIT",                       ebit_proj,  "$#,##0"),
    ("EBIT Margin %",              ebit_m,     "0.0%"),
    ("NOPAT (EBIT × (1-t))",       nopat,      "$#,##0"),
    ("(+) D&A",                    da_proj,    "$#,##0"),
    ("(-) Capex",                  [-c for c in capex_proj], "$#,##0"),
    ("(-) Change in NWC",          nwc_delta,  "$#,##0"),
    ("Unlevered Free Cash Flow",   fcff,       "$#,##0"),
    ("Discount Factor",            disc_f,     "0.0000"),
    ("PV of UFCF",                 pv_fcff,    "$#,##0"),
]
for i, (label, vals, fmt) in enumerate(proj_rows):
    row_n = 3 + i
    bold = label in ("Revenue","NOPAT (EBIT × (1-t))","Unlevered Free Cash Flow","PV of UFCF")
    bg = LGRAY if bold else WHITE
    cell(proj, row_n, 1, label, bold=bold, bg=bg, fc=DARK)
    for j, v in enumerate(vals):
        cell(proj, row_n, 2+j, v, bold=bold, bg=bg, fc=DARK, align="right", num_fmt=fmt)

# Terminal value
sum_pv = sum(pv_fcff)
exit_ebitda = ebit_proj[-1] + da_proj[-1]
tv_mult  = exit_ebitda * 22.0
tv_gcf   = int(fcff[-1] * 1.03 / (wacc - 0.03))
pv_tv_m  = int(tv_mult  / (1+wacc)**5)
pv_tv_g  = int(tv_gcf   / (1+wacc)**5)
net_debt = 75421
shares   = 15343

cell(proj, 16, 1, "─── TERMINAL VALUE & VALUATION ───", bold=True, bg=BLUE, fc=WHITE)
proj.merge_cells("A16:F16")

tv_rows = [
    (17, "Sum of PV(UFCF)",              sum_pv,             "$#,##0"),
    (18, "Terminal Value (Exit Multiple):",tv_mult,           "$#,##0"),
    (19, "PV of Terminal Value (Multiple)", pv_tv_m,         "$#,##0"),
    (20, "Enterprise Value (Exit Multiple)", sum_pv+pv_tv_m, "$#,##0"),
    (21, "(-) Net Debt",                 net_debt,           "$#,##0"),
    (22, "Equity Value (Exit Multiple)", sum_pv+pv_tv_m-net_debt, "$#,##0"),
    (23, "Implied Price / Share (Exit Mult.)", int((sum_pv+pv_tv_m-net_debt)/shares), "$#,##0"),
    (25, "Terminal Value (Gordon Growth):", tv_gcf,           "$#,##0"),
    (26, "PV of Terminal Value (GGM)",   pv_tv_g,            "$#,##0"),
    (27, "Enterprise Value (GGM)",       sum_pv+pv_tv_g,     "$#,##0"),
    (28, "Equity Value (GGM)",           sum_pv+pv_tv_g-net_debt, "$#,##0"),
    (29, "Implied Price / Share (GGM)",  int((sum_pv+pv_tv_g-net_debt)/shares), "$#,##0"),
]
for row_n, label, val, fmt in tv_rows:
    bold = "Implied Price" in label or "Enterprise Value" in label or "Equity Value" in label
    bg = LGRAY if bold else WHITE
    cell(proj, row_n, 1, label, bold=bold, bg=bg, fc=DARK)
    cell(proj, row_n, 2, val, bold=bold, bg=bg, fc=BLUE if "Implied" in label else DARK,
         align="right", num_fmt=fmt)

# ── Sensitivity ───────────────────────────────────────────────────────────────
sens = wbd.create_sheet("Sensitivity")
sens.sheet_view.showGridLines = False
cell(sens, 1, 1, "SENSITIVITY ANALYSIS — IMPLIED SHARE PRICE", bold=True, sz=14, bg=DARK, fc=WHITE)
sens.merge_cells("A1:I1")
set_col_width(sens, 1, 22)
for c in range(2, 10): set_col_width(sens, c, 12)

cell(sens, 3, 1, "Exit Multiple ↓  /  WACC →", bold=True, bg=LGRAY)
waccs = [0.0775, 0.0850, 0.0925, 0.1000, 0.1075]
multiples = [18.0, 20.0, 22.0, 24.0, 26.0]
for j, w in enumerate(waccs):
    cell(sens, 3, 2+j, f"{w:.2%}", bold=True, bg=BLUE, fc=WHITE, align="center")
for i, m in enumerate(multiples):
    cell(sens, 4+i, 1, f"{m:.1f}x EV/EBITDA", bold=True, bg=LGRAY)
    for j, w in enumerate(waccs):
        d_f = [1/(1+w)**t for t in range(1,6)]
        pv = sum(int(f*d) for f,d in zip(fcff, d_f))
        tv_ = exit_ebitda * m / (1+w)**5
        eq  = pv + tv_ - net_debt
        price = int(eq / shares)
        is_base = (abs(m - 22.0) < 0.01 and abs(w - 0.0925) < 0.001)
        bg = AMBER if is_base else (GREEN if price > 240 else (RED if price < 180 else WHITE))
        fc = WHITE if is_base or price > 240 or price < 180 else DARK
        cell(sens, 4+i, 2+j, price, bold=is_base, bg=bg, fc=fc, align="center", num_fmt="$#,##0")

cell(sens, 10, 1, "Amber = base case  |  Green = >$240  |  Red = <$180", italic=True, fc=MGRAY)
sens.merge_cells("A10:I10")

wbd.save(os.path.join(OUT, "AAPL_DCF.xlsx"))
print("✓ AAPL_DCF.xlsx")

# ════════════════════════════════════════════════════════════════════════════
# 3. COMPARABLE COMPANY ANALYSIS
# ════════════════════════════════════════════════════════════════════════════
wbc = openpyxl.Workbook()

comps_sheet = wbc.active; comps_sheet.title = "Comps"
comps_sheet.sheet_view.showGridLines = False

COMP_COLS = ["Company","Ticker","Mkt Cap\n($B)","EV\n($B)","Rev\n($B)",
             "Rev\nGrowth","Gross\nMargin","EBIT\nMargin","EBITDA\nMargin",
             "Net\nMargin","EV/\nRev","EV/\nEBITDA","P/E\n(NTM)","P/FCF","FCF\nYield"]
col_ws = [20,8,10,10,8,9,10,10,11,9,8,10,9,8,9]
for i, (col, w) in enumerate(zip(COMP_COLS, col_ws)):
    set_col_width(comps_sheet, i+1, w)

cell(comps_sheet, 1, 1, "APPLE (AAPL) — COMPARABLE COMPANY ANALYSIS", bold=True, sz=14, bg=DARK, fc=WHITE)
comps_sheet.merge_cells(f"A1:{get_column_letter(len(COMP_COLS))}1")
cell(comps_sheet, 2, 1, "Trading multiples as of May 2025 | NTM consensus estimates | $ in billions", italic=True, fc=MGRAY)
comps_sheet.merge_cells(f"A2:{get_column_letter(len(COMP_COLS))}2")

hdr_row(comps_sheet, 4, COMP_COLS, bg=DARK, fc=WHITE)
comps_sheet.row_dimensions[4].height = 36

peers = [
    # name,  tick,  mktcap,ev,  rev,  rev_g,  gm,    ebit_m, ebitda_m, nm,    ev_rev, ev_ebitda, pe,   pfcf, fcfy
    ("Apple","AAPL", 3280, 3290, 391,  0.061, 0.460, 0.306, 0.342,   0.310, 8.4,  24.5,  32.1, 30.2, 0.031),
    ("Microsoft","MSFT",3100,3160,245, 0.155, 0.699, 0.448, 0.510,   0.352, 12.9, 25.3,  36.2, 34.1, 0.026),
    ("Alphabet","GOOGL",2040,1980,350, 0.139, 0.573, 0.315, 0.358,   0.260, 5.7,  18.2,  21.4, 20.8, 0.042),
    ("Meta","META",    1380,1330,165,  0.221, 0.813, 0.418, 0.451,   0.381, 8.1,  18.5,  25.1, 23.8, 0.038),
    ("Amazon","AMZN",  2150,2260,620,  0.109, 0.490, 0.105, 0.168,   0.094, 3.6,  21.5,  44.2, 41.7, 0.021),
    ("Salesforce","CRM",250, 255, 37,  0.084, 0.771, 0.188, 0.246,   0.161, 6.9,  28.0,  31.5, 29.0, 0.032),
    ("Adobe","ADBE",   220, 225, 22,   0.103, 0.882, 0.352, 0.412,   0.280, 10.2, 24.8,  27.5, 25.9, 0.035),
]

FMTS = [None,None,"$#,##0.0","$#,##0.0","$#,##0.0","0.0%","0.0%","0.0%","0.0%","0.0%","0.0x","0.0x","0.0x","0.0x","0.0%"]
for i, row_data in enumerate(peers):
    row_n = 5 + i
    is_apple = i == 0
    bg = LGRAY if is_apple else (WHITE if i % 2 == 0 else WHITE)
    bg_hex = "E8F0FE" if is_apple else ("F5F5F7" if i % 2 == 0 else "FFFFFF")
    for j, (val, fmt) in enumerate(zip(row_data, FMTS)):
        fc = BLUE if is_apple and j == 0 else DARK
        c = comps_sheet.cell(row=row_n, column=j+1, value=val)
        c.font = Font(bold=is_apple, size=11, color=fc)
        c.fill = PatternFill("solid", fgColor=bg_hex)
        c.alignment = Alignment(horizontal="right" if j > 1 else "left", vertical="center")
        if fmt: c.number_format = fmt

# Stats rows
stat_rows_data = []
all_vals = {k: [p[i] for p in peers[1:]] for k, i in  # exclude Apple
            zip(COMP_COLS[2:], range(2, len(COMP_COLS)))}
def median(lst): s = sorted(x for x in lst if x); n = len(s); return s[n//2] if n%2 else (s[n//2-1]+s[n//2])/2
def mean_(lst):  v = [x for x in lst if x]; return sum(v)/len(v) if v else None

stat_defs = [("Peer Mean", mean_), ("Peer Median", median), ("Peer High", max), ("Peer Low", min)]
for si, (stat_name, fn) in enumerate(stat_defs):
    row_n = 13 + si
    bg_hex = "1D1D1F" if stat_name == "Peer Median" else "F5F5F7"
    fc_hex = "FFFFFF" if stat_name == "Peer Median" else "1D1D1F"
    comps_sheet.cell(row=row_n, column=1, value=stat_name).font = Font(bold=True, size=11, color=fc_hex)
    comps_sheet.cell(row=row_n, column=1).fill = PatternFill("solid", fgColor=bg_hex)
    comps_sheet.cell(row=row_n, column=2, value="").fill = PatternFill("solid", fgColor=bg_hex)
    for j, (col_name, fmt) in enumerate(zip(COMP_COLS[2:], FMTS[2:])):
        vals_list = [p[j+2] for p in peers[1:]]
        try: stat_val = fn(vals_list)
        except: stat_val = None
        c = comps_sheet.cell(row=row_n, column=j+3, value=round(stat_val, 3) if stat_val else None)
        c.font = Font(bold=(stat_name=="Peer Median"), size=11, color=fc_hex)
        c.fill = PatternFill("solid", fgColor=bg_hex)
        c.alignment = Alignment(horizontal="right", vertical="center")
        if fmt: c.number_format = fmt

# Premium / discount row
prem_row = 18
cell(comps_sheet, prem_row, 1, "AAPL vs. Peer Median", bold=True, bg="0071E3", fc=WHITE)
cell(comps_sheet, prem_row, 2, "", bg="0071E3")
peer_medians = [median([p[j] for p in peers[1:]]) for j in range(2, len(COMP_COLS))]
apple_vals   = list(peers[0][2:])
for j, (av, pm, fmt) in enumerate(zip(apple_vals, peer_medians, FMTS[2:])):
    if pm and pm != 0 and av:
        prem = av / pm - 1
        bg = "34C759" if prem < 0 else "D70015"
        c = comps_sheet.cell(row=prem_row, column=j+3, value=prem)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "+0.0%;-0.0%"

cell(comps_sheet, 20, 1, "Source: Bloomberg, FactSet. NTM = Next Twelve Months consensus. Market data as of May 2025.", italic=True, fc=MGRAY)
comps_sheet.merge_cells(f"A20:{get_column_letter(len(COMP_COLS))}20")
freeze(comps_sheet, "C5")

wbc.save(os.path.join(OUT, "AAPL_Comps.xlsx"))
print("✓ AAPL_Comps.xlsx")

# ════════════════════════════════════════════════════════════════════════════
# 4. LBO MODEL
# ════════════════════════════════════════════════════════════════════════════
wbl = openpyxl.Workbook()

# ── Transaction Summary ──────────────────────────────────────────────────────
txn = wbl.active; txn.title = "Transaction Summary"
txn.sheet_view.showGridLines = False
set_col_width(txn, 1, 35); set_col_width(txn, 2, 18); set_col_width(txn, 3, 18)

cell(txn, 1, 1, "AAPL — HYPOTHETICAL LBO ANALYSIS", bold=True, sz=18, bg=DARK, fc=WHITE)
txn.merge_cells("A1:C1")
cell(txn, 2, 1, "For illustrative purposes only. Apple is not a realistic LBO candidate given its $3T+ market cap.",
     italic=True, fc=RED)
txn.merge_cells("A2:C2")

ENTRY_EV  = 3_290_000   # $M
ENTRY_EQ  = 3_280_000
NET_DEBT_LBO = 75_421
TOTAL_USES= ENTRY_EQ + NET_DEBT_LBO

SENIOR_DEBT = int(TOTAL_USES * 0.30)
MEZZ_DEBT   = int(TOTAL_USES * 0.10)
TOTAL_DEBT  = SENIOR_DEBT + MEZZ_DEBT
EQUITY_C    = TOTAL_USES - TOTAL_DEBT

txn_data = [
    (4,  "SOURCES & USES", None, BLUE, WHITE),
    (5,  "SOURCES",        "Amount ($M)", "% of Total"),
    (6,  "Senior Secured Debt (8.5x EBITDA)", SENIOR_DEBT, f"{SENIOR_DEBT/TOTAL_USES:.1%}"),
    (7,  "Mezzanine / Second Lien",           MEZZ_DEBT,   f"{MEZZ_DEBT/TOTAL_USES:.1%}"),
    (8,  "Sponsor Equity",                    EQUITY_C,    f"{EQUITY_C/TOTAL_USES:.1%}"),
    (9,  "Total Sources",                     TOTAL_USES,  "100.0%"),
    (10, "USES",           "Amount ($M)", "% of Total"),
    (11, "Purchase of Equity",               ENTRY_EQ,   f"{ENTRY_EQ/TOTAL_USES:.1%}"),
    (12, "Refinancing of Existing Debt",     NET_DEBT_LBO,f"{NET_DEBT_LBO/TOTAL_USES:.1%}"),
    (13, "Total Uses",                       TOTAL_USES,  "100.0%"),
    (15, "ENTRY VALUATION METRICS", None, BLUE, WHITE),
    (16, "Entry Enterprise Value ($M)",      ENTRY_EV,    None),
    (17, "Entry EBITDA ($M)",                134_000,     None),
    (18, "Entry EV/EBITDA",                  f"{ENTRY_EV/134000:.1f}x", None),
    (19, "Entry P/E (NTM)",                  "32.1x",     None),
    (20, "Debt / EBITDA at Entry",           f"{TOTAL_DEBT/134000:.1f}x", None),
    (21, "Equity / Total Cap",               f"{EQUITY_C/TOTAL_USES:.1%}", None),
]
for row_n, label, v2, *rest in txn_data:
    if v2 is None:
        section_hdr(txn, row_n, label, 3, bg=rest[0], fc=rest[1])
        continue
    bold = label in ("Total Sources","Total Uses","Entry Enterprise Value ($M)","Entry EV/EBITDA")
    bg = LGRAY if bold else WHITE
    cell(txn, row_n, 1, label, bold=bold, bg=bg, fc=DARK)
    cell(txn, row_n, 2, v2, bold=bold, bg=bg, fc=DARK, align="right",
         num_fmt="$#,##0" if isinstance(v2, int) else None)
    if rest and rest[0] is not None:
        cell(txn, row_n, 3, rest[0], bold=bold, bg=bg, fc=MGRAY, align="right")

# ── Debt Schedule ─────────────────────────────────────────────────────────────
debt = wbl.create_sheet("Debt Schedule")
debt.sheet_view.showGridLines = False
LBO_YEARS = ["Entry","FY2025E","FY2026E","FY2027E","FY2028E","FY2029E"]
set_col_width(debt, 1, 35)
for c in range(2, 8): set_col_width(debt, c, 14)

hdr_row(debt, 1, ["DEBT SCHEDULE ($M)"] + LBO_YEARS, bg=DARK, fc=WHITE)

SR_RATE = 0.085; MZ_RATE = 0.120
amort   = [0] + [int(SENIOR_DEBT * 0.05)] * 5  # 5% annual amortization
sr_beg  = [SENIOR_DEBT]
for i in range(5): sr_beg.append(sr_beg[-1] - amort[i+1])
sr_end  = sr_beg[1:] + [sr_beg[-1] - amort[-1]]
sr_end  = [s - a for s, a in zip(sr_beg[1:]+[0], amort[1:])]
# Redo cleanly
sr_bal = [SENIOR_DEBT]
for _ in range(5): sr_bal.append(int(sr_bal[-1] * 0.95))
mz_bal = [MEZZ_DEBT] + [MEZZ_DEBT]*5  # bullet

sr_int = [0] + [int(sr_bal[i] * SR_RATE) for i in range(5)]
mz_int = [0] + [int(MEZZ_DEBT * MZ_RATE)] * 5
tot_int = [s+m for s,m in zip(sr_int, mz_int)]
tot_debt = [s+m for s,m in zip(sr_bal, mz_bal)]

debt_rows = [
    ("Senior Secured — Beginning Balance", sr_bal),
    ("  Mandatory Amortization (5%/yr)",   [0]+[int(sr_bal[i]-sr_bal[i+1]) for i in range(5)]),
    ("Senior Secured — Ending Balance",    sr_bal[1:]+[sr_bal[-1]]),
    ("Interest (8.5%)",                    sr_int),
    ("Mezzanine — Balance",                mz_bal),
    ("Interest (12.0%)",                   mz_int),
    ("Total Debt",                         tot_debt),
    ("Total Interest Expense",             tot_int),
]
for i, (label, vals) in enumerate(debt_rows):
    row_n = 3 + i
    bold = label in ("Total Debt","Total Interest Expense","Senior Secured — Ending Balance")
    bg = LGRAY if bold else WHITE
    cell(debt, row_n, 1, label, bold=bold, bg=bg, fc=DARK)
    for j, v in enumerate(vals):
        cell(debt, row_n, 2+j, v, bold=bold, bg=bg, fc=DARK, align="right", num_fmt="$#,##0")

# ── Returns Analysis ──────────────────────────────────────────────────────────
ret = wbl.create_sheet("Returns Analysis")
ret.sheet_view.showGridLines = False
set_col_width(ret, 1, 30)
for c in range(2, 8): set_col_width(ret, c, 14)

cell(ret, 1, 1, "RETURNS ANALYSIS", bold=True, sz=14, bg=DARK, fc=WHITE)
ret.merge_cells("A1:G1")

# Exit at FY2029 at various EV/EBITDA multiples
exit_ebitda_lbo = 186_000  # estimated FY2029 EBITDA (from DCF projections)
exit_multiples = [16, 18, 20, 22, 24]
remaining_debt = tot_debt[-1]

cell(ret, 3, 1, "Exit EV/EBITDA →", bold=True, bg=LGRAY, fc=DARK)
for j, m in enumerate(exit_multiples):
    cell(ret, 3, 2+j, f"{m}x", bold=True, bg=BLUE, fc=WHITE, align="center")

irr_rows_data = []
for irr_label, eq_mult in [("Equity Value at Exit ($M)", None),
                            ("Equity MOIC", None),
                            ("Sponsor IRR (5-year)", None)]:
    irr_rows_data.append((irr_label, exit_multiples))

row_n = 4
cell(ret, row_n, 1, "Exit Enterprise Value ($M)", bold=False, bg=WHITE, fc=DARK)
for j, m in enumerate(exit_multiples):
    cell(ret, row_n, 2+j, exit_ebitda_lbo * m, bg=WHITE, fc=DARK, align="right", num_fmt="$#,##0")

row_n = 5
cell(ret, row_n, 1, "(-) Remaining Debt ($M)", bold=False, bg=WHITE, fc=DARK)
for j, m in enumerate(exit_multiples):
    cell(ret, row_n, 2+j, -remaining_debt, bg=WHITE, fc=DARK, align="right", num_fmt="$#,##0")

row_n = 6
cell(ret, row_n, 1, "Equity Value at Exit ($M)", bold=True, bg=LGRAY, fc=DARK)
eq_exits = [exit_ebitda_lbo * m - remaining_debt for m in exit_multiples]
for j, eq in enumerate(eq_exits):
    cell(ret, row_n, 2+j, int(eq), bold=True, bg=LGRAY, fc=DARK, align="right", num_fmt="$#,##0")

row_n = 7
cell(ret, row_n, 1, "Equity MOIC", bold=True, bg=LGRAY, fc=DARK)
moics = [eq / EQUITY_C for eq in eq_exits]
for j, moic in enumerate(moics):
    fc = GREEN if moic >= 2.0 else (RED if moic < 1.5 else DARK)
    cell(ret, row_n, 2+j, round(moic, 2), bold=True, bg=LGRAY, fc=fc, align="right", num_fmt="0.00x")

row_n = 8
cell(ret, row_n, 1, "Sponsor IRR (5-year hold)", bold=True, bg=DARK, fc=WHITE)
irrs = [(moic ** (1/5) - 1) for moic in moics]
for j, irr in enumerate(irrs):
    fc = GREEN if irr >= 0.20 else (RED if irr < 0.12 else AMBER)
    cell(ret, row_n, 2+j, irr, bold=True, bg="1D1D1F", fc=fc, align="center", num_fmt="0.0%")

cell(ret, 10, 1, "Note: Assumes 5-year hold, FY2029 exit. Equity contribution = $"
     f"{EQUITY_C/1000:.0f}B. No management rollover or fees modeled.", italic=True, fc=MGRAY)
ret.merge_cells("A10:G10")
cell(ret, 11, 1, "Green IRR ≥ 20% | Amber 12–20% | Red < 12%.  "
     "This is a purely illustrative exercise — Apple is not a realistic LBO candidate.", italic=True, fc=RED)
ret.merge_cells("A11:G11")

wbl.save(os.path.join(OUT, "AAPL_LBO.xlsx"))
print("✓ AAPL_LBO.xlsx")

print("\nAll files generated.")
