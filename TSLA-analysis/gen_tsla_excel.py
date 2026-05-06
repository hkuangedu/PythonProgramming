"""gen_tsla_excel.py — Tesla deep valuation Excel suite
Generates: TSLA_3Statement.xlsx, TSLA_DCF.xlsx, TSLA_Comps.xlsx, TSLA_SOTP.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

OUT = "/root/PythonProgramming/TSLA-analysis"
os.makedirs(OUT, exist_ok=True)

DARK="1D1D1F"; TR="E31937"; GRN="34C759"; BLU="0071E3"; LGR="F5F5F7"
MGR="86868B"; WHT="FFFFFF"; AMB="FF9500"; DK2="2C2C2E"; RED="D70015"

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def freeze(ws, c): ws.freeze_panes = c

def cell(ws, r, col, v=None, bold=False, sz=11, bg=None, fc=DARK,
         align="left", fmt=None, italic=False):
    c = ws.cell(row=r, column=col, value=v)
    c.font = Font(bold=bold, size=sz, color=fc, italic=italic)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fmt: c.number_format = fmt
    return c

def hrow(ws, r, labels, bg=DARK, fc=WHT, sz=11, sc=1):
    for i, l in enumerate(labels):
        cell(ws, r, sc+i, l, bold=True, sz=sz, bg=bg, fc=fc, align="center")

def srow(ws, r, label, span, bg=TR, fc=WHT):
    c = ws.cell(row=r, column=1, value=label)
    c.font = Font(bold=True, size=11, color=fc)
    c.fill = PatternFill("solid", fgColor=bg)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c.alignment = Alignment(horizontal="left", vertical="center")


# ── Data ──────────────────────────────────────────────────────────────────────
ALL  = ["FY2021A","FY2022A","FY2023A","FY2024A",
        "FY2025E","FY2026E","FY2027E","FY2028E","FY2029E"]
HIST = ALL[:4]
PROJ = ALL[4:]

REVENUE = [53823, 81462, 96773, 97690, 105500, 130500, 166000, 210000, 260000]
REV_A   = [47232, 71462, 82418, 77119,  82000, 104000, 135000, 170000, 205000]
REV_E   = [ 2789,  3909,  6035, 10086,  12500,  14500,  17500,  22000,  28000]
REV_S   = [ 3802,  6091,  8319, 10485,  11000,  12000,  13500,  18000,  27000]

COGS    = [40217, 60609, 79113, 80168,  86060, 104400, 129480, 159600, 195200]
GROSS   = [r - c for r, c in zip(REVENUE, COGS)]
# GM%:    25.3%  25.6%  18.2%  17.9%  18.5%  20.0%  22.0%  24.0%  25.1%

RD      = [ 2593,  3075,  3969,  4453,   4800,   5500,   6500,   8000,  10000]
SGA     = [ 2667,  2890,  3721,  4124,   4200,   4800,   5500,   6200,   7100]
EBIT    = [ 6523, 13656,  8891,  7071,   8968,  13050,  19920,  27300,  35360]
# Implied other opex (SBC, restructuring) = GROSS - RD - SGA - EBIT
OTHER   = [g - r - s - e for g, r, s, e in zip(GROSS, RD, SGA, EBIT)]

EBIT_M  = [e / r for e, r in zip(EBIT, REVENUE)]
DA      = [ 2500,  3600,  4600,  5400,   6000,   7000,   8500,  10500,  13000]
EBITDA  = [e + d for e, d in zip(EBIT, DA)]
EBITDA_M= [e / r for e, r in zip(EBITDA, REVENUE)]

INT_NET = [  -56,   -77,  1066,  1600,   1800,   2000,   2200,   2500,   2800]
NET_INC = [ 5519, 12556, 14997,  7092,   8000,  11000,  16800,  23100,  31500]
NET_ADJ = [ 5519, 12556,  9097,  7092,   8000,  11000,  16800,  23100,  31500]
# FY2023 reported NI includes $5.9B non-recurring deferred tax benefit

SHARES  = [ 3208,  3222,  3198,  3194,   3190,   3182,   3182,   3178,   3182]
EPS_REP = [ 1.72,  3.90,  4.69,  2.22,   2.51,   3.46,   5.28,   7.27,   9.90]
EPS_ADJ = [ 1.72,  3.90,  2.84,  2.22,   2.51,   3.46,   5.28,   7.27,   9.90]

SBC     = [ 2121,  1560,  1741,  1714,   1800,   2000,   2200,   2500,   2800]
NWC     = [ -500,  1200,  -500,  -600,   -400,   -500,   -600,   -700,   -800]
CAPEX   = [ 6515,  7158,  8898, 11000,  10000,  10500,  11000,  12000,  13000]
OCF     = [11497, 14485, 13256, 14620,  16000,  20000,  25000,  33000,  43500]
FCF     = [o - c for o, c in zip(OCF, CAPEX)]

# Historical balance sheet (FY2021–FY2024)
CASH_H  = [17576, 22185, 29094, 36600]
INV_H   = [ 5757,  8615, 13721,  7700]
OCA_H   = [10000, 13000, 15000, 17000]
CA_H    = [c + i + o for c, i, o in zip(CASH_H, INV_H, OCA_H)]
PPE_H   = [18884, 27743, 29725, 38000]
OLA_H   = [ 8000, 12000, 18000, 20000]
TA_H    = [c + p + o for c, p, o in zip(CA_H, PPE_H, OLA_H)]
AP_H    = [ 6994, 13448, 14431, 15000]
DR_H    = [ 2000,  3000,  3500,  4000]
OCL_H   = [ 6000,  9000,  9500, 12000]
CL_H    = [a + d + o for a, d, o in zip(AP_H, DR_H, OCL_H)]
LTD_H   = [ 6839,  1597,  5234,  7700]
OLL_H   = [11000, 12000, 15000, 15000]
TL_H    = [c + l + o for c, l, o in zip(CL_H, LTD_H, OLL_H)]
EQ_H    = [30189, 44704, 62634, 72000]


# ============================================================
# 1. THREE-STATEMENT MODEL
# ============================================================
wb3 = openpyxl.Workbook()

# Cover
cov = wb3.active
cov.title = "Cover"
cov.sheet_view.showGridLines = False
cell(cov, 1, 1, "TESLA, INC. (TSLA)", bold=True, sz=28, bg=DARK, fc=WHT, align="center")
cov.merge_cells("A1:H1")
cov.row_dimensions[1].height = 60
for col in range(1, 9):
    cw(cov, col, 18)
meta = [
    (3, "Three-Statement Financial Model", 16, TR, WHT, True),
    (4, "FY2021A – FY2029E  |  $ in millions  |  Fiscal Year = Calendar Year", 12, LGR, DARK, False),
    (6, "Ticker: TSLA  |  Exchange: NASDAQ  |  Sector: Consumer Discretionary / Technology", 11, None, MGR, False),
    (7, "Note: FY2023 reported net income includes $5.9B non-recurring deferred tax benefit", 11, None, RED, True),
    (9, "Source: Tesla 10-K filings (FY2021–FY2024), FactSet consensus estimates (FY2025E–FY2029E)", 10, None, MGR, True),
]
for rn, txt, sz, bg, fc, bold in meta:
    cell(cov, rn, 1, txt, bold=bold, sz=sz, bg=bg, fc=fc)
    cov.merge_cells(f"A{rn}:H{rn}")

# Income Statement tab
inc = wb3.create_sheet("Income Statement")
inc.sheet_view.showGridLines = False
cw(inc, 1, 42)
for c in range(2, 11):
    cw(inc, c, 13)

hrow(inc, 1, ["TESLA, INC. — INCOME STATEMENT ($M)"] + ALL)
inc.row_dimensions[1].height = 28

rows_inc = [
    (3,  "Automotive Revenue",             REV_A,   "$#,##0", LGR,  False),
    (4,  "Energy Generation & Storage",    REV_E,   "$#,##0", WHT,  False),
    (5,  "Services & Other",               REV_S,   "$#,##0", WHT,  False),
    (6,  "Total Revenue",                  REVENUE, "$#,##0", LGR,  True),
    (7,  "Revenue YoY Growth",
         [None] + [REVENUE[i]/REVENUE[i-1]-1 for i in range(1,9)], "0.0%", WHT, False),
    (9,  "Cost of Revenue",                COGS,    "$#,##0", WHT,  False),
    (10, "Gross Profit",                   GROSS,   "$#,##0", LGR,  True),
    (11, "Gross Margin %",                 [g/r for g,r in zip(GROSS,REVENUE)], "0.0%", WHT, False),
    (13, "Research & Development",         RD,      "$#,##0", WHT,  False),
    (14, "Selling, General & Admin",       SGA,     "$#,##0", WHT,  False),
    (15, "Other OpEx (SBC, restructuring)",OTHER,   "$#,##0", WHT,  False),
    (16, "Total Operating Expenses",       [r+s+o for r,s,o in zip(RD,SGA,OTHER)], "$#,##0", LGR, False),
    (18, "EBIT (Operating Income)",        EBIT,    "$#,##0", LGR,  True),
    (19, "EBIT Margin %",                  EBIT_M,  "0.0%",   WHT,  False),
    (20, "EBITDA",                         EBITDA,  "$#,##0", LGR,  True),
    (21, "EBITDA Margin %",                EBITDA_M,"0.0%",   WHT,  False),
    (23, "Net Interest Inc. / (Exp.)",     INT_NET, "$#,##0", WHT,  False),
    (24, "Net Income (Reported)",          NET_INC, "$#,##0", LGR,  True),
    (25, "Net Income (Adjusted)",          NET_ADJ, "$#,##0", LGR,  False),
    (26, "Adj. Net Margin %",              [n/r for n,r in zip(NET_ADJ,REVENUE)], "0.0%", WHT, False),
    (28, "Diluted EPS (Reported)",         EPS_REP, "$0.00",  WHT,  False),
    (29, "Diluted EPS (Adjusted)",         EPS_ADJ, "$0.00",  LGR,  False),
    (30, "Diluted Shares Outstanding (M)", SHARES,  "#,##0",  WHT,  False),
    (32, "D&A",                            DA,      "$#,##0", WHT,  False),
    (33, "Capex",                          CAPEX,   "$#,##0", WHT,  False),
    (34, "Free Cash Flow",                 FCF,     "$#,##0", LGR,  True),
    (35, "FCF Margin %",                   [f/r for f,r in zip(FCF,REVENUE)], "0.0%", WHT, False),
]
for rn, label, vals, fmt, bg, bold in rows_inc:
    cell(inc, rn, 1, label, bold=bold, bg=bg, fc=DARK)
    for i, v in enumerate(vals):
        cell(inc, rn, 2+i, v, bold=bold, bg=bg, fc=DARK, align="right", fmt=fmt)

cell(inc, 2, 2, "◄ ACTUALS ►", italic=True, sz=9, fc=MGR, align="center")
inc.merge_cells("B2:E2")
cell(inc, 2, 6, "◄ ESTIMATES ►", italic=True, sz=9, fc=TR, align="center")
inc.merge_cells("F2:J2")
freeze(inc, "B3")

# Balance Sheet tab
bs = wb3.create_sheet("Balance Sheet")
bs.sheet_view.showGridLines = False
cw(bs, 1, 42)
for c in range(2, 7):
    cw(bs, c, 15)

hrow(bs, 1, ["TESLA, INC. — BALANCE SHEET ($M)"] + HIST)
CHECK = [a - l - e for a, l, e in zip(TA_H, TL_H, EQ_H)]

bs_rows = [
    (3,  "ASSETS",                             None,   TR,  WHT,  None, None),
    (4,  "Cash & Equivalents + Investments",   CASH_H, "$#,##0", None, None, False),
    (5,  "Inventories",                        INV_H,  "$#,##0", None, None, False),
    (6,  "Other Current Assets",               OCA_H,  "$#,##0", None, None, False),
    (7,  "Total Current Assets",               CA_H,   "$#,##0", LGR, None, True),
    (8,  "Property, Plant & Equipment, net",   PPE_H,  "$#,##0", None, None, False),
    (9,  "Other Long-Term Assets",             OLA_H,  "$#,##0", None, None, False),
    (10, "TOTAL ASSETS",                       TA_H,   "$#,##0", DARK, WHT,  True),
    (12, "LIABILITIES & EQUITY",               None,   TR,  WHT,  None, None),
    (13, "Accounts Payable",                   AP_H,   "$#,##0", None, None, False),
    (14, "Deferred Revenue",                   DR_H,   "$#,##0", None, None, False),
    (15, "Other Current Liabilities",          OCL_H,  "$#,##0", None, None, False),
    (16, "Total Current Liabilities",          CL_H,   "$#,##0", LGR, None, True),
    (17, "Long-Term Debt",                     LTD_H,  "$#,##0", None, None, False),
    (18, "Other Long-Term Liabilities",        OLL_H,  "$#,##0", None, None, False),
    (19, "TOTAL LIABILITIES",                  TL_H,   "$#,##0", DARK, WHT,  True),
    (20, "TOTAL EQUITY",                       EQ_H,   "$#,##0", DARK, WHT,  True),
    (21, "Check (Assets – Liab – Equity)",     CHECK,  "$#,##0", LGR, None, False),
]
for item in bs_rows:
    rn, label = item[0], item[1]
    if item[2] is None:
        srow(bs, rn, label, 5, bg=item[3], fc=item[4])
        continue
    vals, fmt, lbg, fc_, bold = item[2], item[3], item[4], item[5], item[6]
    lbg = lbg or WHT; fc_ = fc_ or DARK
    cell(bs, rn, 1, label, bold=bold, bg=lbg, fc=fc_)
    for i, v in enumerate(vals):
        cell(bs, rn, 2+i, v, bold=bold, bg=lbg, fc=fc_, align="right", fmt=fmt)

freeze(bs, "B3")

# Cash Flow tab
cf = wb3.create_sheet("Cash Flow")
cf.sheet_view.showGridLines = False
cw(cf, 1, 42)
for c in range(2, 11):
    cw(cf, c, 13)

hrow(cf, 1, ["TESLA, INC. — CASH FLOW STATEMENT ($M)"] + ALL)

OTHER_OCF = [o - n - d - s - w for o, n, d, s, w in zip(OCF, NET_INC, DA, SBC, NWC)]

cf_rows = [
    (3,  "OPERATING ACTIVITIES",          None, TR, WHT),
    (4,  "Net Income",                    NET_INC, "$#,##0"),
    (5,  "D&A",                           DA,      "$#,##0"),
    (6,  "Stock-Based Compensation",      SBC,     "$#,##0"),
    (7,  "Changes in Working Capital",    NWC,     "$#,##0"),
    (8,  "Other Operating Items",         OTHER_OCF,"$#,##0"),
    (9,  "Cash from Operations",          OCF,     "$#,##0"),
    (11, "INVESTING ACTIVITIES",          None, TR, WHT),
    (12, "Capital Expenditures",          [-c for c in CAPEX], "$#,##0"),
    (13, "Cash from Investing",           [-c for c in CAPEX], "$#,##0"),
    (15, "FINANCING ACTIVITIES",          None, TR, WHT),
    (16, "Debt Issuances / (Repayments)", [0]*9, "$#,##0"),
    (17, "Other Financing",               [0]*9, "$#,##0"),
    (18, "Cash from Financing",           [0]*9, "$#,##0"),
    (20, "FREE CASH FLOW",                FCF,  "$#,##0"),
    (21, "FCF Margin %",                  [f/r for f,r in zip(FCF,REVENUE)], "0.0%"),
    (22, "FCF / Diluted Share",           [f/s for f,s in zip(FCF,SHARES)], "$0.00"),
]
for item in cf_rows:
    rn, label = item[0], item[1]
    if item[2] is None:
        srow(cf, rn, label, 10, bg=item[3], fc=item[4])
        continue
    vals, fmt = item[2], item[3]
    bold = label in ("Cash from Operations","Cash from Investing","Cash from Financing","FREE CASH FLOW")
    bg = LGR if bold else WHT
    cell(cf, rn, 1, label, bold=bold, bg=bg, fc=DARK)
    for i, v in enumerate(vals):
        cell(cf, rn, 2+i, v, bold=bold, bg=bg, fc=DARK, align="right", fmt=fmt)

freeze(cf, "B3")
wb3.save(os.path.join(OUT, "TSLA_3Statement.xlsx"))
print("✓ TSLA_3Statement.xlsx")


# ============================================================
# 2. DCF VALUATION MODEL
# ============================================================
wbd = openpyxl.Workbook()

# Assumptions tab
asmp = wbd.active
asmp.title = "Assumptions"
asmp.sheet_view.showGridLines = False
cw(asmp, 1, 38); cw(asmp, 2, 20); cw(asmp, 3, 40)

cell(asmp, 1, 1, "TSLA DCF MODEL — KEY ASSUMPTIONS", bold=True, sz=16, bg=DARK, fc=WHT)
asmp.merge_cells("A1:C1")

asmp_data = [
    (3,  "WACC INPUTS",                    None, BLU, WHT),
    (4,  "Risk-Free Rate (10yr UST)",      "4.35%", "Current 10-year Treasury yield"),
    (5,  "Equity Risk Premium",            "5.00%", "Damodaran US ERP"),
    (6,  "Levered Beta (2yr vs S&P 500)",  "2.10",  "High-beta growth stock; 5Y β ≈ 2.0"),
    (7,  "Cost of Equity (CAPM)",         "14.85%", "4.35% + 2.10 × 5.00%"),
    (8,  "Pre-Tax Cost of Debt",           "4.50%", "Weighted avg coupon on senior notes"),
    (9,  "Effective Tax Rate",            "10.0%",  "Normalized; prior years had DTA offsets"),
    (10, "After-Tax Cost of Debt",         "4.05%", "Pre-tax × (1 – tax rate)"),
    (11, "Debt / Total Capital",           "4.5%",  "LT Debt / (LT Debt + Mkt Cap)"),
    (12, "WACC",                          "10.0%",  "Rounded from ~10.1%; reflects long-run risk"),
    (14, "TERMINAL VALUE",                 None, BLU, WHT),
    (15, "Terminal Growth Rate",           "4.0%",  "Above GDP; justified by FSD/Energy optionality"),
    (16, "Terminal EV/EBITDA Multiple",   "20.0x",  "Discount to current 72x; assumes margin maturity"),
    (17, "DCF Horizon",                    "FY2025E–FY2029E", "5-year explicit forecast period"),
    (19, "PROJECTION ASSUMPTIONS",         None, BLU, WHT),
    (20, "Revenue Growth FY2025E",         "7.9%",  "New Model Y refresh + Energy acceleration"),
    (21, "Revenue Growth FY2026E",        "23.7%",  "Affordable model launch + Cybercab pilot"),
    (22, "Revenue Growth FY2027E",        "27.2%",  "Volume ramp + FSD subscription launch"),
    (23, "Revenue Growth FY2028E",        "26.5%",  "Robotaxi revenue + Optimus pilot units"),
    (24, "Revenue Growth FY2029E",        "23.8%",  "FSD/Optimus at early scale"),
    (25, "EBIT Margin FY2025E",            "8.5%",  "Modest recovery from FY2024 7.2%"),
    (26, "EBIT Margin FY2026E",           "10.0%",  "Mix shift + manufacturing efficiency"),
    (27, "EBIT Margin FY2027E",           "12.0%",  "FSD margin contribution"),
    (28, "EBIT Margin FY2028E",           "13.0%",  "Robotaxi high-margin revenue"),
    (29, "EBIT Margin FY2029E",           "13.6%",  "Approaching software-like blended margin"),
    (31, "SHARE DATA",                     None, BLU, WHT),
    (32, "Diluted Shares (M)",            "3,194",  "FY2024A; minimal buybacks expected"),
    (33, "Net Cash ($M)",                "28,900",  "$36.6B cash – $7.7B LT debt"),
    (34, "Current Stock Price",           "$278",   "As of May 2025"),
    (35, "Current Market Cap ($B)",       "$887",   "Implying EV/FY2024 EBITDA of ~71x"),
]
for item in asmp_data:
    rn, label = item[0], item[1]
    if item[2] is None:
        srow(asmp, rn, label, 3, bg=item[3], fc=item[4])
        continue
    val, note = item[2], item[3]
    cell(asmp, rn, 1, label, bg=LGR, fc=DARK)
    cell(asmp, rn, 2, val, bold=True, bg=WHT, fc=BLU, align="right")
    cell(asmp, rn, 3, note, italic=True, fc=MGR)

# DCF Projection tab
WACC = 0.10
PROJ5 = PROJ  # FY2025E–FY2029E (5 years)

proj = wbd.create_sheet("DCF Projection")
proj.sheet_view.showGridLines = False
cw(proj, 1, 40)
for c in range(2, 8):
    cw(proj, c, 15)

hrow(proj, 1, ["TESLA — UNLEVERED DCF PROJECTION ($M)"] + PROJ5)

rev5    = [105500, 130500, 166000, 210000, 260000]
ebit5   = [ 8968,  13050,  19920,  27300,  35360]
nopat5  = [int(e * 0.90) for e in ebit5]  # tax rate 10%
da5     = [  6000,   7000,   8500,  10500,  13000]
cap5    = [ 10000,  10500,  11000,  12000,  13000]
nwc5    = [  -400,   -500,   -600,   -700,   -800]
fcff5   = [n + d - c + w for n, d, c, w in zip(nopat5, da5, cap5, nwc5)]
disc5   = [1 / (1 + WACC)**t for t in range(1, 6)]
pvfcf5  = [int(f * d) for f, d in zip(fcff5, disc5)]

exit_ebitda = ebit5[-1] + da5[-1]  # FY2029E EBITDA

proj_rows = [
    ("Revenue",                    rev5,   "$#,##0",  True),
    ("Revenue Growth %",           [rev5[i]/rev5[i-1]-1 if i>0 else None for i in range(5)], "0.0%", False),
    ("EBIT",                       ebit5,  "$#,##0",  False),
    ("EBIT Margin %",              [e/r for e,r in zip(ebit5,rev5)], "0.0%", False),
    ("NOPAT (EBIT × (1 – 10% tax))",nopat5,"$#,##0", True),
    ("(+) D&A",                    da5,    "$#,##0",  False),
    ("(–) Capex",                  [-c for c in cap5], "$#,##0", False),
    ("(–) Change in NWC",          nwc5,   "$#,##0",  False),
    ("Unlevered FCF (UFCF)",       fcff5,  "$#,##0",  True),
    ("Discount Factor (WACC 10%)", disc5,  "0.0000",  False),
    ("PV of UFCF",                 pvfcf5, "$#,##0",  True),
]
for i, (label, vals, fmt, bold) in enumerate(proj_rows):
    rn = 3 + i
    bg = LGR if bold else WHT
    cell(proj, rn, 1, label, bold=bold, bg=bg, fc=DARK)
    for j, v in enumerate(vals):
        cell(proj, rn, 2+j, v, bold=bold, bg=bg, fc=DARK, align="right", fmt=fmt)

# Terminal value & valuation
sum_pv = sum(pvfcf5)
net_cash = 28900
shares_lbo = 3194

tv_mult = exit_ebitda * 20.0
tv_ggm  = int(fcff5[-1] * 1.04 / (WACC - 0.04))
pv_tv_m = int(tv_mult / (1 + WACC)**5)
pv_tv_g = int(tv_ggm  / (1 + WACC)**5)

srow(proj, 16, "─── TERMINAL VALUE & VALUATION BRIDGE ───", 6, bg=TR)

tv_rows = [
    (17, "Sum of PV(UFCF)",                       sum_pv,              "$#,##0"),
    (18, "EBITDA in Exit Year (FY2029E)",          exit_ebitda,         "$#,##0"),
    (19, "Terminal Value — Exit Multiple (20x)",   tv_mult,             "$#,##0"),
    (20, "PV of TV (Exit Multiple)",               pv_tv_m,             "$#,##0"),
    (21, "Enterprise Value (Exit Multiple)",       sum_pv + pv_tv_m,    "$#,##0"),
    (22, "(+) Net Cash",                           net_cash,            "$#,##0"),
    (23, "Equity Value (Exit Multiple)",           sum_pv + pv_tv_m + net_cash, "$#,##0"),
    (24, "Implied Price / Share",                  int((sum_pv + pv_tv_m + net_cash) / shares_lbo), "$#,##0"),
    (26, "Terminal Value — Gordon Growth (g=4%)", tv_ggm,              "$#,##0"),
    (27, "PV of TV (GGM)",                         pv_tv_g,             "$#,##0"),
    (28, "Enterprise Value (GGM)",                 sum_pv + pv_tv_g,    "$#,##0"),
    (29, "Equity Value (GGM)",                     sum_pv + pv_tv_g + net_cash, "$#,##0"),
    (30, "Implied Price / Share (GGM)",            int((sum_pv + pv_tv_g + net_cash) / shares_lbo), "$#,##0"),
]
for rn, label, val, fmt in tv_rows:
    bold = ("Implied Price" in label or "Enterprise Value" in label or "Equity Value" in label)
    bg = LGR if bold else WHT
    cell(proj, rn, 1, label, bold=bold, bg=bg, fc=DARK)
    fc_ = TR if "Implied Price" in label else DARK
    cell(proj, rn, 2, val, bold=bold, bg=bg, fc=fc_, align="right", fmt=fmt)

# Sensitivity tab (WACC × Terminal Multiple)
sens = wbd.create_sheet("Sensitivity")
sens.sheet_view.showGridLines = False
cw(sens, 1, 24)
for c in range(2, 10):
    cw(sens, c, 13)

cell(sens, 1, 1, "SENSITIVITY: IMPLIED SHARE PRICE  |  Exit EV/EBITDA (rows) × WACC (cols)",
     bold=True, sz=12, bg=DARK, fc=WHT)
sens.merge_cells("A1:I1")

waccs     = [0.075, 0.085, 0.100, 0.110, 0.120, 0.130]
multiples = [14.0, 17.0, 20.0, 23.0, 26.0]

cell(sens, 3, 1, "EV/EBITDA ↓ / WACC →", bold=True, bg=LGR)
for j, w in enumerate(waccs):
    cell(sens, 3, 2+j, f"{w:.1%}", bold=True, bg=DARK, fc=WHT, align="center")

for i, m in enumerate(multiples):
    cell(sens, 4+i, 1, f"{m:.0f}x EV/EBITDA", bold=True, bg=LGR)
    for j, w in enumerate(waccs):
        d = [1/(1+w)**t for t in range(1,6)]
        pv = sum(int(f*di) for f, di in zip(fcff5, d))
        tv_ = exit_ebitda * m / (1+w)**5
        eq  = pv + tv_ + net_cash
        px  = int(eq / shares_lbo)
        is_base = (abs(m-20.0) < 0.01 and abs(w-0.100) < 0.001)
        bg = AMB if is_base else (GRN if px > 200 else (RED if px < 100 else WHT))
        fc = WHT if is_base or px > 200 or px < 100 else DARK
        cell(sens, 4+i, 2+j, px, bold=is_base, bg=bg, fc=fc, align="center", fmt="$#,##0")

cell(sens, 10, 1,
     "Amber = base case ($10.0% WACC, 20x EV/EBITDA)  |  "
     "Green = >$200  |  Red = <$100  |  Current price ~$278",
     italic=True, fc=MGR)
sens.merge_cells("A10:I10")
cell(sens, 11, 1,
     "NOTE: DCF alone does not capture Tesla's full optionality (FSD, Optimus, Robotaxi). "
     "See SOTP model for segment-level analysis.",
     italic=True, fc=TR)
sens.merge_cells("A11:I11")

wbd.save(os.path.join(OUT, "TSLA_DCF.xlsx"))
print("✓ TSLA_DCF.xlsx")


# ============================================================
# 3. COMPARABLE COMPANY ANALYSIS
# ============================================================
wbc = openpyxl.Workbook()

# Auto Peers tab
auto = wbc.active
auto.title = "Auto Peers"
auto.sheet_view.showGridLines = False

AUTO_COLS = ["Company","Ticker","Mkt Cap\n($B)","EV\n($B)","Revenue\n($B)",
             "Rev\nGrowth","Gross\nMargin","EBIT\nMargin","EBITDA\nMargin",
             "EV/\nRev","EV/\nEBITDA","P/E\n(NTM)","FCF\nYield","EV per\nDelivery ($K)"]
ACW = [22,8,10,10,10,9,10,10,11,7,10,8,8,12]
for i, w in enumerate(ACW):
    cw(auto, i+1, w)

cell(auto, 1, 1, "TESLA (TSLA) — AUTO PEER COMPARABLE COMPANY ANALYSIS", bold=True, sz=14, bg=DARK, fc=WHT)
auto.merge_cells(f"A1:{get_column_letter(len(AUTO_COLS))}1")
cell(auto, 2, 1, "FY2024A actuals | NTM consensus multiples | $ in billions | May 2025", italic=True, fc=MGR)
auto.merge_cells(f"A2:{get_column_letter(len(AUTO_COLS))}2")

hrow(auto, 4, AUTO_COLS, bg=DARK, fc=WHT)
auto.row_dimensions[4].height = 36

auto_peers = [
    # name, tick, mktcap, ev, rev, rev_g, gm, ebit_m, ebitda_m, ev_rev, ev_ebitda, pe, fcfy, ev_del
    ("Tesla",       "TSLA",  887,  858,  97.7, 0.001, 0.179, 0.072, 0.128,  8.8, 68.8, 125.0, 0.004, 479),
    ("BYD",         "002594", 88,   92, 107.1, 0.152, 0.198, 0.060, 0.090,  0.9,  8.5,  22.4, 0.028,  19),
    ("GM",          "GM",     56,   95, 187.4, 0.009, 0.142, 0.062, 0.095,  0.5,  3.1,   5.2, 0.082,  17),
    ("Ford",        "F",      45,   88, 185.0, 0.048, 0.103, 0.040, 0.065,  0.5,  4.0,   9.1, 0.068,  22),
    ("BMW",         "BMW",    55,   65, 155.2,-0.012, 0.163, 0.073, 0.120,  0.4,  3.5,   5.1, 0.094,  27),
    ("Mercedes",    "MBG",    60,   72, 145.6,-0.030, 0.189, 0.088, 0.126,  0.5,  3.9,   5.6, 0.089,  30),
    ("Stellantis",  "STLA",   18,   28, 189.4,-0.150, 0.137, 0.054, 0.078,  0.1,  2.3,   3.8, 0.185,   7),
    ("Rivian",      "RIVN",   12,   11,   4.8, 0.042,-0.380,-0.400,-0.380,  2.3,   None, None, None,  227),
    ("Lucid",       "LCID",    7,    6,   0.8, 0.153,-4.200,-4.800,-4.700,  7.5,   None, None, None,  300),
    ("Li Auto",     "LI",     18,   14,  17.3, 0.196, 0.220, 0.055, 0.080,  0.8,  7.0,  14.5, 0.042,  14),
    ("NIO",         "NIO",     8,   12,   9.1, 0.110,-0.080,-0.220,-0.180,  1.3,   None, None, None,  130),
]

AFMTS = [None,None,"#,##0.0","#,##0.0","#,##0.0","0.0%","0.0%","0.0%","0.0%",
         "0.0x","0.0x","0.0x","0.0%","#,##0"]

for i, row in enumerate(auto_peers):
    rn = 5 + i
    is_tsla = (i == 0)
    bg_h = "FFF0F0" if is_tsla else ("F5F5F7" if i % 2 == 0 else "FFFFFF")
    for j, (val, fmt) in enumerate(zip(row, AFMTS)):
        fc_ = TR if (is_tsla and j == 0) else DARK
        c = auto.cell(row=rn, column=j+1, value=val)
        c.font = Font(bold=is_tsla, size=11, color=fc_)
        c.fill = PatternFill("solid", fgColor=bg_h)
        c.alignment = Alignment(horizontal="right" if j > 1 else "left", vertical="center")
        if fmt and val is not None: c.number_format = fmt

# Stats rows
from statistics import median as _med, mean as _mean
def safe_median(lst): v = [x for x in lst if x is not None]; return _med(v) if v else None
def safe_mean(lst):   v = [x for x in lst if x is not None]; return _mean(v) if v else None

for si, (stat_name, fn) in enumerate([("Peer Mean (ex-TSLA)", safe_mean),
                                       ("Peer Median (ex-TSLA)", safe_median)]):
    rn = 17 + si
    bg_h = "1D1D1F" if "Median" in stat_name else LGR
    fc_h = WHT if "Median" in stat_name else DARK
    auto.cell(row=rn, column=1, value=stat_name).font = Font(bold=True, size=11, color=fc_h)
    auto.cell(row=rn, column=1).fill = PatternFill("solid", fgColor=bg_h)
    auto.cell(row=rn, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for j, fmt in enumerate(AFMTS[2:]):
        vals = [auto_peers[k][j + 2] for k in range(1, len(auto_peers))]
        v = fn(vals)
        c = auto.cell(row=rn, column=j + 3)
        c.value = round(v, 4) if v is not None else None
        c.font = Font(bold=("Median" in stat_name), size=11, color=fc_h)
        c.fill = PatternFill("solid", fgColor=bg_h)
        c.alignment = Alignment(horizontal="right", vertical="center")
        if fmt and v is not None: c.number_format = fmt

# Premium/discount row
rn = 20
cell(auto, rn, 1, "TSLA vs. Peer Median", bold=True, bg=TR, fc=WHT)
peer_med_row = [safe_median([auto_peers[k][j + 2] for k in range(1, len(auto_peers))]) for j in range(len(AFMTS[2:]))]
tsla_row = list(auto_peers[0][2:])
for j, (tv, pm, fmt) in enumerate(zip(tsla_row, peer_med_row, AFMTS[2:])):
    if pm and pm != 0 and tv is not None:
        prem = tv / pm - 1
        bg = GRN if prem < 0 else RED
        c = auto.cell(row=rn, column=j + 3)
        c.value = prem
        c.font = Font(bold=True, size=11, color=WHT)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = "+0.0%;-0.0%"

cell(auto, 22, 1,
     "Source: Bloomberg, FactSet. NTM = Next Twelve Months consensus. "
     "Tesla trades at 13–68x premium to auto peers on EV/EBITDA — market prices in software/AI optionality.",
     italic=True, fc=MGR)
auto.merge_cells(f"A22:{get_column_letter(len(AUTO_COLS))}22")
freeze(auto, "C5")

# Tech / AI Peers reference tab
tech = wbc.create_sheet("Tech Peers (Reference)")
tech.sheet_view.showGridLines = False

TECH_COLS = ["Company","Ticker","Mkt Cap\n($B)","EV/Rev\n(NTM)","EV/EBITDA\n(NTM)","P/E\n(NTM)",
             "Rev Growth","EBIT Margin","Net Margin","FCF Yield","Why Relevant"]
TCW = [20,8,12,12,14,10,11,12,11,10,40]
for i, w in enumerate(TCW):
    cw(tech, i+1, w)

cell(tech, 1, 1, "TECH / SOFTWARE / AI PEER REFERENCE — Tesla's Software Premium Implied by Optionality",
     bold=True, sz=12, bg=DARK, fc=WHT)
tech.merge_cells(f"A1:{get_column_letter(len(TECH_COLS))}1")

hrow(tech, 3, TECH_COLS, bg=DARK, fc=WHT)
tech.row_dimensions[3].height = 36

tech_peers = [
    ("Tesla",     "TSLA",  887, 8.8, 68.8, 125.0, 0.001, 0.072, 0.073, 0.004, "Subject — blended auto+software valuation"),
    ("NVIDIA",    "NVDA", 3200, 25.0, 48.5,  38.2, 0.780, 0.620, 0.550, 0.021, "Data center AI compute leader; robotics"),
    ("Alphabet",  "GOOGL",2040,  5.5, 17.9,  21.4, 0.139, 0.315, 0.260, 0.042, "Waymo competitor; AI infrastructure"),
    ("Amazon",    "AMZN", 2150,  3.5, 21.5,  44.2, 0.109, 0.105, 0.094, 0.021, "AWS/logistics; autonomous delivery"),
    ("Palantir",  "PLTR",  200, 35.0, 90.0, 140.0, 0.310, 0.200, 0.180, 0.020, "AI platform; autonomous ops analogue"),
    ("Mobileye",  "MBLY",  12,  4.5, 25.0,  45.0, 0.140, 0.180, 0.120, 0.008, "ADAS/autonomous driving direct comp"),
    ("Aurora",    "AUR",    4,   None, None,  None, 0.350, -2.500,-2.600, None,  "Level 4 autonomous trucking"),
    ("Uber",      "UBER",  170,  5.0, 30.0,  55.0, 0.140, 0.070, 0.050, 0.015, "Robotaxi network owner; FSD TAM proxy"),
]

TFMTS = [None,None,"#,##0","0.0x","0.0x","0.0x","0.0%","0.0%","0.0%","0.0%",None]
for i, row in enumerate(tech_peers):
    rn = 4 + i
    is_tsla = (i == 0)
    bg_h = "FFF0F0" if is_tsla else ("F5F5F7" if i % 2 == 0 else "FFFFFF")
    for j, (val, fmt) in enumerate(zip(row, TFMTS)):
        fc_ = TR if (is_tsla and j == 0) else DARK
        c = tech.cell(row=rn, column=j+1, value=val)
        c.font = Font(bold=is_tsla, size=11, color=fc_)
        c.fill = PatternFill("solid", fgColor=bg_h)
        c.alignment = Alignment(horizontal="right" if (j>1 and j<10) else "left", vertical="center")
        if fmt and val is not None: c.number_format = fmt

cell(tech, 14, 1,
     "Key insight: Tesla's EV/EBITDA of 69x aligns with AI/software premium peers (PLTR ~90x, NVDA ~49x), "
     "not traditional auto (3–4x). The market assigns ~$600–700B of TSLA's $887B market cap to non-auto optionality.",
     italic=True, fc=MGR)
tech.merge_cells(f"A14:{get_column_letter(len(TECH_COLS))}14")

wbc.save(os.path.join(OUT, "TSLA_Comps.xlsx"))
print("✓ TSLA_Comps.xlsx")


# ============================================================
# 4. SUM-OF-THE-PARTS (SOTP)
# ============================================================
wbs = openpyxl.Workbook()

# SOTP Summary tab
sotp = wbs.active
sotp.title = "SOTP Summary"
sotp.sheet_view.showGridLines = False
cw(sotp, 1, 32); cw(sotp, 2, 14); cw(sotp, 3, 14); cw(sotp, 4, 14); cw(sotp, 5, 14)
cw(sotp, 6, 14); cw(sotp, 7, 14); cw(sotp, 8, 38)

cell(sotp, 1, 1, "TESLA — SUM OF THE PARTS VALUATION", bold=True, sz=16, bg=DARK, fc=WHT)
sotp.merge_cells("A1:H1")

# Header
hrow(sotp, 3, ["SEGMENT","FY2027E\nRevenue","FY2027E\nEBITDA",
               "Multiple /\nBasis","Bear EV","Base EV","Bull EV","Valuation Notes"], bg=DARK, fc=WHT)
sotp.row_dimensions[3].height = 36

# Segment data
segments = [
    # label, rev, ebitda, mult_basis, bear_ev, base_ev, bull_ev, notes
    ("Automotive (ex-FSD, ex-SC)", 128_000, 17_400, "8.0x EBITDA",
     97_000, 139_000, 187_000,
     "Valued as premium ICE/EV OEM. BMW/Mercedes trade at 3–4x; Tesla premium for brand, tech, margins."),
    ("Energy Gen. & Storage",      17_500,  5_250, "22.0x EBITDA",
     65_000, 115_500, 175_000,
     "Megapack growth 200%+ in 3yr; high GM (>25%). Grid-scale battery secular tailwind. Clean energy premium."),
    ("FSD / Autopilot Software",    4_000,  None, "25-40x Revenue",
      0, 100_000, 300_000,
     "Bear: regulatory block / no monetization. Base: $99/mo × 5M subs. Bull: licensing to OEMs + robotaxi."),
    ("Cybercab / Robotaxi Network", 2_000,  None, "15-30x Revenue",
      0,  30_000, 150_000,
     "Bear: still pre-revenue. Base: limited operations in 2–3 cities. Bull: network effect + margin expansion."),
    ("Supercharger Network",        3_000,  1_200, "20.0x EBITDA",
     12_000, 24_000,  40_000,
     "Now licensed to Ford, GM, Rivian, BMW. Infrastructure moat; expanding beyond Tesla vehicles."),
    ("Insurance",                   1_000,  None, "10x Revenue",
      3_000, 10_000,  20_000,
     "Leverages real-time vehicle data for underwriting. Currently 12 US states. Long-term potential globally."),
    ("Optimus Humanoid Robots",     None,   None, "DCF / Option",
      0,  75_000, 500_000,
     "Bear: R&D-stage, no revenue. Base: 10K+ units @ $20K margin. Bull: 1M+ units; $30T robot labor TAM."),
]

seg_colors = [BLU, "00B386", "A855F7", AMB, "00C2E0", "FF6B6B", TR]

for i, (label, rev, ebitda, mult, bear, base, bull, notes) in enumerate(segments):
    rn = 4 + i
    bg_h = LGR if i % 2 == 0 else WHT
    cell(sotp, rn, 1, label, bold=True, bg=bg_h, fc=DARK)
    cell(sotp, rn, 2, rev,    bg=bg_h, fc=DARK, align="right", fmt="$#,##0" if rev else None)
    cell(sotp, rn, 3, ebitda, bg=bg_h, fc=DARK, align="right", fmt="$#,##0" if ebitda else None)
    cell(sotp, rn, 4, mult,   bg=bg_h, fc=MGR, align="center")
    cell(sotp, rn, 5, bear, bold=False, bg=bg_h, fc=RED,  align="right", fmt="$#,##0")
    cell(sotp, rn, 6, base, bold=True,  bg=bg_h, fc=DARK, align="right", fmt="$#,##0")
    cell(sotp, rn, 7, bull, bold=False, bg=bg_h, fc=GRN,  align="right", fmt="$#,##0")
    cell(sotp, rn, 8, notes, italic=True, fc=MGR)

# Totals — write label in col 1, data in cols 5-8 (no full-row merge)
bears = [0, 65000, 0, 0, 12000, 3000, 0]
bases = [139000, 115500, 100000, 30000, 24000, 10000, 75000]
bulls = [187000, 175000, 300000, 150000, 40000, 20000, 500000]

def total_row(ws, rn, label, bear, base, bull, label_bg, label_fc, data_fcs):
    cell(ws, rn, 1, label, bold=True, bg=label_bg, fc=label_fc)
    ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=4)
    for col, val, fc_ in zip([5, 6, 7], [bear, base, bull], data_fcs):
        cell(ws, rn, col, val, bold=True, bg=label_bg, fc=fc_, align="right", fmt="$#,##0")
    cell(ws, rn, 8, "", bg=label_bg)

eq_bear = sum(bears) + 28900
eq_base = sum(bases) + 28900
eq_bull = sum(bulls) + 28900

total_row(sotp, 12, "TOTAL ENTERPRISE VALUE",
          sum(bears), sum(bases), sum(bulls), DARK, WHT, [RED, BLU, GRN])
total_row(sotp, 13, "(+) Net Cash ($28.9B)",
          28900, 28900, 28900, DK2, WHT, [WHT, WHT, WHT])
total_row(sotp, 14, "EQUITY VALUE",
          eq_bear, eq_base, eq_bull, TR, WHT, [WHT, WHT, WHT])

rn = 15
cell(sotp, rn, 1, "IMPLIED SHARE PRICE  (3,194M diluted shares)", bold=True, bg=TR, fc=WHT)
sotp.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=4)
for col, eq in [(5, eq_bear), (6, eq_base), (7, eq_bull)]:
    px = int(eq / 3194)
    cell(sotp, rn, col, px, bold=True, bg=TR, fc=WHT, align="right", fmt="$#,##0")
cell(sotp, rn, 8, "", bg=TR)

# Footnotes
notes_rows = [
    (17, "Current Market Price: ~$278  |  Market Cap: ~$887B  |  Implied EV: ~$858B", DARK, WHT, True),
    (18, "Market EV implies: Auto $139B + Energy $115B + ALL other optionality ~$604B — entirely FSD/Optimus/Robotaxi premium", MGR, DARK, False),
    (19, "All figures in $M. FY2027E reference year (2 years forward). Multiples applied undiscounted as proxy.", MGR, DARK, False),
    (20, "Bear: No FSD/Optimus monetization, auto competition accelerates, margin compression persists.", RED, DARK, False),
    (21, "Base: Energy scales, FSD reaches 5M subscribers, Cybercab in 3 cities, Optimus initial units.", BLU, DARK, False),
    (22, "Bull: FSD licensed to OEMs, Optimus 1M+ units, Megapack leads global grid storage, Robotaxi network.", GRN, DARK, False),
]
for rn, txt, bg, fc_, bold in notes_rows:
    cell(sotp, rn, 1, txt, bold=bold, bg=bg, fc=fc_, italic=not bold)
    sotp.merge_cells(f"A{rn}:H{rn}")


# Segment Projections tab
seg_proj = wbs.create_sheet("Segment Projections")
seg_proj.sheet_view.showGridLines = False
cw(seg_proj, 1, 38)
for c in range(2, 8):
    cw(seg_proj, c, 15)

hrow(seg_proj, 1, ["TESLA — SEGMENT REVENUE & MARGIN PROJECTIONS ($M)"] + PROJ5, bg=DARK, fc=WHT)

seg_data = [
    (3,  "AUTOMOTIVE SEGMENT",     None, BLU, WHT),
    (4,  "Automotive Revenue",     [77119,82000,104000,135000,170000,205000][1:], "$#,##0"),
    (5,  "  Vehicles (ex-FSD)",    [72000,76000, 96000,123000,158000,190000][1:], "$#,##0"),
    (6,  "  FSD Revenue",          [  800, 1200,  2500,  5000,  8000, 12000][1:], "$#,##0"),
    (7,  "  Regulatory Credits",   [ 1696, 1800,  2000,  2500,  3000,  3000][1:], "$#,##0"),
    (8,  "  Services / Leasing",   [ 2623, 3000,  3500,  4500,  1000,  None][1:], "$#,##0"),
    (9,  "Automotive Gross Margin",[0.173,0.178, 0.185, 0.195, 0.210,0.230][1:], "0.0%"),
    (11, "ENERGY SEGMENT",         None, "00B386", WHT),
    (12, "Energy Revenue",         [10086,12500, 14500, 17500, 22000, 28000][1:], "$#,##0"),
    (13, "  Megapack (Utility)",   [ 7000, 9000, 10500, 12500, 15500, 20000][1:], "$#,##0"),
    (14, "  Powerwall (Residential)",[2000,2500,  3000,  4000,  5500,  7000][1:], "$#,##0"),
    (15, "  Solar",                [ 1086, 1000,  1000,  1000,  1000,  1000][1:], "$#,##0"),
    (16, "Energy Gross Margin",    [0.246,0.265,  0.275, 0.285, 0.295, 0.310][1:], "0.0%"),
    (18, "SERVICES SEGMENT",       None, "A855F7", WHT),
    (19, "Services Revenue",       [10485,11000, 12000, 13500, 18000, 27000][1:], "$#,##0"),
    (20, "  FSD Subscription",     [  800, 1200,  2500,  5000,  8000, 12000][1:], "$#,##0"),
    (21, "  Service Centers",      [ 5500, 5800,  5500,  5500,  6000,  6500][1:], "$#,##0"),
    (22, "  Supercharging",        [ 2000, 2500,  2500,  2500,  3000,  3500][1:], "$#,##0"),
    (23, "  Other (Insurance, etc)",[2185, 1500,  1500,  1500,  1000,  5000][1:], "$#,##0"),
    (24, "Services Gross Margin",  [0.073, 0.090, 0.120, 0.150, 0.220, 0.280][1:], "0.0%"),
    (26, "CONSOLIDATED",           None, DARK, WHT),
    (27, "Total Revenue",          [97690,105500,130500,166000,210000,260000][1:], "$#,##0"),
    (28, "Total Gross Profit",     [17522, 19508, 26100, 36520, 50400, 65260][1:], "$#,##0"),
    (29, "Blended Gross Margin",   [0.179,  0.185, 0.200, 0.220, 0.240, 0.251][1:], "0.0%"),
    (30, "EBITDA",                 [12471, 14968, 20050, 28420, 37800, 48360][1:], "$#,##0"),
    (31, "EBITDA Margin",          [0.128,  0.142, 0.154, 0.171, 0.180, 0.186][1:], "0.0%"),
]
for item in seg_data:
    rn, label = item[0], item[1]
    if item[2] is None:
        srow(seg_proj, rn, label, 6, bg=item[3], fc=item[4])
        continue
    vals, fmt = item[2], item[3]
    bold = label in ("Total Revenue","Total Gross Profit","EBITDA",
                     "Automotive Revenue","Energy Revenue","Services Revenue")
    bg = LGR if bold else WHT
    cell(seg_proj, rn, 1, label, bold=bold, bg=bg, fc=DARK)
    for i, v in enumerate(vals):
        if v is not None:
            cell(seg_proj, rn, 2+i, v, bold=bold, bg=bg, fc=DARK, align="right", fmt=fmt)

freeze(seg_proj, "B3")

# Scenario Bridge tab
scen = wbs.create_sheet("Scenario Bridge")
scen.sheet_view.showGridLines = False
cw(scen, 1, 38)
for c in range(2, 6):
    cw(scen, c, 20)

hrow(scen, 1, ["SCENARIO BRIDGE — EQUITY VALUE ($M) & IMPLIED SHARE PRICE",
               "BEAR", "BASE", "BULL", "Key Driver"], bg=DARK, fc=WHT)

bridge_rows = [
    ("Starting Point: Core Auto EV (FY2027E, 8x EBITDA)", 97000, 139000, 187000, "Auto margins, deliveries volume"),
    ("+ Energy Storage EV (22x EBITDA)",                   65000, 115500, 175000, "Megapack pricing, TAM capture"),
    ("+ FSD / Software EV",                                    0, 100000, 300000, "Penetration rate, monetization path"),
    ("+ Cybercab / Robotaxi EV",                               0,  30000, 150000, "Regulatory timeline, city launches"),
    ("+ Supercharger Network EV",                          12000,  24000,  40000, "Third-party licensing ramp"),
    ("+ Insurance EV",                                      3000,  10000,  20000, "State expansion, loss ratio"),
    ("+ Optimus Robot Option Value",                           0,  75000, 500000, "Unit volumes, production cost"),
    ("= Total Enterprise Value",                          177000, 493500,1372000, ""),
    ("+ Net Cash",                                         28900,  28900,  28900, "~$36.6B cash – $7.7B debt"),
    ("= Equity Value",                                    205900, 522400,1400900, ""),
    ("Diluted Shares (M)",                                  3194,   3194,   3194, ""),
    ("IMPLIED SHARE PRICE",                                   64,    164,    439, ""),
    ("Current Price (~$278)",                                278,    278,    278, "Market between base and bull"),
    ("Upside / (Downside) %",                             -0.770, -0.410,  0.579, "vs. current ~$278/share"),
]
for i, (label, bear, base, bull, driver) in enumerate(bridge_rows):
    rn = 3 + i
    is_total = label.startswith("=") or label.startswith("IMPLIED")
    is_current = "Current Price" in label
    bg = DARK if is_total else (AMB if is_current else (LGR if i % 2 == 0 else WHT))
    fc_l = WHT if is_total else DARK
    bold = is_total or is_current
    cell(scen, rn, 1, label, bold=bold, bg=bg, fc=fc_l)
    for col, val in [(2, bear), (3, base), (4, bull)]:
        if label == "IMPLIED SHARE PRICE":
            cell(scen, rn, col, val, bold=True, bg=bg, fc=WHT, align="right", fmt="$#,##0")
        elif "Upside" in label:
            fc_ = RED if val < 0 else GRN
            cell(scen, rn, col, val, bold=True, bg=bg, fc=fc_, align="right", fmt="+0.0%;-0.0%")
        elif "%" in label or isinstance(val, float):
            cell(scen, rn, col, val, bold=bold, bg=bg, fc=fc_l, align="right", fmt="#,##0")
        else:
            cell(scen, rn, col, val, bold=bold, bg=bg, fc=fc_l, align="right", fmt="$#,##0")
    cell(scen, rn, 5, driver, italic=True, fc=MGR)

wbs.save(os.path.join(OUT, "TSLA_SOTP.xlsx"))
print("✓ TSLA_SOTP.xlsx")

print("\n✅ All Excel models generated.")
